from __future__ import annotations

import datetime as dt
from typing import Final, cast

from openpyxl import Workbook
from openpyxl.cell.cell import Cell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from src.calendar.utils import is_holiday_or_weekend
from src.domain.context import VarKey
from src.domain.types import ShiftType, Weekday

SEPARATOR: Final[str] = ","  # セル内で複数人を区切る文字列


def export_schedule_to_excel(
    *,
    assignment: dict[VarKey, int],
    shortage_slack: dict[tuple[str, dt.date], float] | None = None,
    days: list[dt.date],
    hospital_names: list[str],
    out_path: str,
) -> None:
    wb = Workbook()

    ws_like = wb.active
    # None/Chartsheet の可能性を潰す
    if not isinstance(ws_like, Worksheet):
        ws_like = wb.create_sheet(title="勤務表")
    ws: Worksheet = cast(Worksheet, ws_like)
    ws.title = "勤務表"

    # 列幅・見た目
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 6
    for idx in range(len(hospital_names)):
        col_letter = get_column_letter(3 + idx)
        ws.column_dimensions[col_letter].width = 10

    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    holiday_fill = PatternFill(fill_type="solid", start_color="FFF4CC", end_color="FFF4CC")
    shortage_fill = PatternFill(fill_type="solid", start_color="FF9999", end_color="FF9999")

    # 見出し
    a1: Cell = ws.cell(row=1, column=1, value="日付")
    b1: Cell = ws.cell(row=1, column=2, value="曜日")
    a1.font = bold
    b1.font = bold
    a1.alignment = center
    b1.alignment = center

    for j, hname in enumerate(hospital_names, start=3):
        cell: Cell = ws.cell(row=1, column=j, value=hname)
        cell.font = bold
        cell.alignment = center

    weekday_list: list[Weekday] = list(Weekday)

    cell_values: dict[tuple[dt.date, str], list[str]] = {}
    for (h, w, d, s), v in assignment.items():
        if v == 0:
            continue
        suffix = "AM" if s == ShiftType.AM else "PM" if s == ShiftType.PM else ""
        label = f"{w}{suffix}".strip()  # 例: "IVR01AM", "診断05PM", "治療02"
        cell_values.setdefault((d, h), []).append(label)

    # 本体行
    for i, d in enumerate(days, start=2):
        c_date = ws.cell(row=i, column=1, value=d.isoformat())
        c_date.alignment = center
        jp_weekday = weekday_list[d.weekday()].value
        c_wd = ws.cell(row=i, column=2, value=jp_weekday)
        c_wd.alignment = center

        for j, hname in enumerate(hospital_names, start=3):
            labels = cell_values.get((d, hname), [])
            txt = SEPARATOR.join(labels)
            cell = ws.cell(row=i, column=j, value=txt)
            cell.alignment = left
            # 人手不足があれば色付け
            if shortage_slack and (hname, d) in shortage_slack:
                cell.fill = shortage_fill

        if is_holiday_or_weekend(d):  # -> bool を返すよう定義側も注釈
            for col in range(1, 3 + len(hospital_names)):
                ws.cell(row=i, column=col).fill = holiday_fill

        for col in range(1, 3 + len(hospital_names)):
            ws.cell(row=i, column=col).border = border

    for col in range(1, 3 + len(hospital_names)):
        ws.cell(row=1, column=col).border = border

    ws.freeze_panes = "C2"
    wb.save(out_path)
