# tests/test_export_excel.py
from __future__ import annotations

import datetime as dt
from pathlib import Path

from openpyxl import load_workbook

from src.domain.types import ShiftType
from src.io.export_excel import export_schedule_to_excel


def _rgb(cell):
    """openpyxl の色は '00FFF4CC' のように先頭にアルファが付くことがあるので末尾6桁で比較"""
    fill = cell.fill
    if not fill or fill.fill_type != "solid":
        return None
    c = fill.start_color
    rgb = getattr(c, "rgb", None)
    if not rgb:
        return None
    return rgb[-6:].upper()  # e.g., 'FFF4CC'


def test_export_excel_basic_grid_and_styles(tmp_path: Path):
    # ==== 準備 ====
    # 2025-10-03 (金), 2025-10-04 (土), 2025-10-05 (日)
    d1 = dt.date(2025, 10, 3)
    d2 = dt.date(2025, 10, 4)  # 土曜 → 休日塗りつぶし期待
    d3 = dt.date(2025, 10, 5)  # 日曜 → 休日塗りつぶし期待
    days = [d1, d2, d3]

    hospitals = ["大学", "病院A"]
    # (h, w, d, s) -> 1/0
    assignment = {
        (hospitals[0], "診断01", d1, ShiftType.DAY): 1,  # サフィックス無し
        (hospitals[0], "診断02", d1, ShiftType.AM): 1,  # " AM"
        (hospitals[1], "診断03", d1, ShiftType.PM): 1,  # " PM"
        (hospitals[1], "診断04", d2, ShiftType.NIGHT): 1,  # サフィックス無し(当直)
        # 同一セルに複数(AM/PM)を入れて連結動作確認
        (hospitals[0], "診断05", d3, ShiftType.AM): 1,
        (hospitals[0], "診断06", d3, ShiftType.PM): 1,
    }

    out = tmp_path / "schedule.xlsx"

    # ==== 実行 ====
    export_schedule_to_excel(
        assignment=assignment,
        days=days,
        hospital_names=hospitals,
        out_path=str(out),
    )

    assert out.exists(), "Excelファイルが作成されていません"

    # ==== 検証 ====
    wb = load_workbook(out)
    ws = wb.active

    # シート名
    assert ws.title == "勤務表"

    # 見出し
    assert ws["A1"].value == "日付"
    assert ws["B1"].value == "曜日"
    assert ws["C1"].value == hospitals[0]
    assert ws["D1"].value == hospitals[1]

    # フリーズペイン
    assert ws.freeze_panes == "C2"

    # 1行目以外のグリッド:A列=日付 / B列=曜 / C..=病院
    # 行のオフセットは2行目から days[0]
    # d1: 2025-10-03(金)
    r1 = 2
    assert ws.cell(r1, 1).value == d1.isoformat()
    assert ws.cell(r1, 2).value in ("金曜", "金")  # 実装は "金曜"、将来の表記変更に少し寛容に
    # 病院セル
    # 大学: 診断01 + 診断02 AM → "診断01 / 診断02 AM" の順とは限らないので集合比較
    c_univ_d1 = ws.cell(r1, 3).value or ""
    parts = {p.strip() for p in c_univ_d1.split(",") if p.strip()}
    assert {"診断01", "診断02AM"} == parts

    # 病院A: 診断03 PM
    assert (ws.cell(r1, 4).value or "").strip() == "診断03PM"

    # d2: 土曜行が薄いオレンジか
    r2 = 3
    assert ws.cell(r2, 1).value == d2.isoformat()
    # 行全体が塗られていること(代表としてA列を検査)
    assert _rgb(ws.cell(r2, 1)) == "FFF4CC"

    # 夜勤はサフィックス無し
    assert (ws.cell(r2, 4).value or "").strip() == "診断04"

    # d3: 日曜行が薄いオレンジ
    r3 = 4
    assert ws.cell(r3, 1).value == d3.isoformat()
    assert _rgb(ws.cell(r3, 1)) == "FFF4CC"

    # 同一セルの連結(AM/PM)
    c_univ_d3 = (ws.cell(r3, 3).value or "").strip()
    parts3 = {p.strip() for p in c_univ_d3.split(",") if p.strip()}
    assert {"診断05AM", "診断06PM"} == parts3


def test_empty_cells_are_blank(tmp_path: Path):
    """割当が無いセルは空文字 or None が入っている(エクスポートで例外が出ない)"""
    d = dt.date(2025, 10, 1)
    days = [d]
    hospitals = ["A病院", "B病院"]
    assignment = {}  # 何も割当無し

    out = tmp_path / "empty.xlsx"
    export_schedule_to_excel(
        assignment=assignment,
        days=days,
        hospital_names=hospitals,
        out_path=str(out),
    )
    wb = load_workbook(out)
    ws = wb.active

    assert ws["A1"].value == "日付"
    assert ws["B1"].value == "曜日"
    assert ws["C1"].value == hospitals[0]
    assert ws["D1"].value == hospitals[1]

    # データ行は空
    r = 2
    assert ws.cell(r, 1).value == d.isoformat()
    assert (ws.cell(r, 3).value or "") == ""
    assert (ws.cell(r, 4).value or "") == ""
